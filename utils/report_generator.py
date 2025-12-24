"""Report generation - Create PDF and HTML reports"""

from datetime import datetime
from typing import Dict
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate PDF, HTML, and Excel reports"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_pdf_report(self, symbol: str, data: Dict) -> str:
        """Generate PDF report"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            
            filename = self.output_dir / f"{symbol}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            doc = SimpleDocTemplate(str(filename), pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title = Paragraph(f"Stock Analysis Report: {symbol}", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Company info
            company_info = f"""
            <b>Company:</b> {data.get('company_name', 'N/A')}<br/>
            <b>Current Price:</b> ${data.get('price', 0):.2f}<br/>
            <b>P/E Ratio:</b> {data.get('pe_ratio', 'N/A')}<br/>
            <b>Market Cap:</b> ${data.get('market_cap', 0):,.0f}<br/>
            """
            story.append(Paragraph(company_info, styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Quality score
            quality_text = f"""
            <b>Quality Score:</b> {data.get('quality_score', 0):.1f}/100<br/>
            <b>Recommendation:</b> {data.get('recommendation', 'N/A')}<br/>
            """
            story.append(Paragraph(quality_text, styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Thesis
            story.append(Paragraph("<b>Investment Thesis:</b>", styles['Heading2']))
            thesis = data.get('thesis', 'N/A')
            story.append(Paragraph(thesis.replace('\n', '<br/>'), styles['Normal']))
            
            # Build PDF
            doc.build(story)
            logger.info(f"PDF report generated: {filename}")
            return str(filename)
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            return ""
    
    def generate_html_report(self, symbol: str, data: Dict) -> str:
        """Generate HTML report"""
        try:
            filename = self.output_dir / f"{symbol}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{symbol} Analysis Report</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 40px;
                        background-color: #f5f5f5;
                    }}
                    .container {{
                        background-color: white;
                        padding: 20px;
                        border-radius: 8px;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    }}
                    h1 {{
                        color: #0d47a1;
                        border-bottom: 2px solid #0d47a1;
                        padding-bottom: 10px;
                    }}
                    h2 {{
                        color: #1565c0;
                        margin-top: 30px;
                    }}
                    .metrics {{
                        display: grid;
                        grid-template-columns: repeat(2, 1fr);
                        gap: 20px;
                        margin: 20px 0;
                    }}
                    .metric {{
                        background-color: #f9f9f9;
                        padding: 15px;
                        border-left: 4px solid #0d47a1;
                    }}
                    .metric-label {{
                        font-weight: bold;
                        color: #0d47a1;
                    }}
                    .metric-value {{
                        font-size: 1.2em;
                        color: #333;
                    }}
                    .score {{
                        font-size: 2em;
                        font-weight: bold;
                        color: #0d47a1;
                        margin: 20px 0;
                    }}
                    .recommendation {{
                        background-color: #e3f2fd;
                        padding: 15px;
                        border-radius: 4px;
                        margin: 20px 0;
                    }}
                    .footer {{
                        margin-top: 40px;
                        padding-top: 20px;
                        border-top: 1px solid #ccc;
                        font-size: 0.9em;
                        color: #666;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>Stock Analysis Report: {symbol}</h1>
                    
                    <div class="metrics">
                        <div class="metric">
                            <div class="metric-label">Current Price</div>
                            <div class="metric-value">${data.get('price', 0):.2f}</div>
                        </div>
                        <div class="metric">
                            <div class="metric-label">P/E Ratio</div>
                            <div class="metric-value">{data.get('pe_ratio', 'N/A')}</div>
                        </div>
                        <div class="metric">
                            <div class="metric-label">52 Week High</div>
                            <div class="metric-value">${data.get('52w_high', 0):.2f}</div>
                        </div>
                        <div class="metric">
                            <div class="metric-label">52 Week Low</div>
                            <div class="metric-value">${data.get('52w_low', 0):.2f}</div>
                        </div>
                    </div>
                    
                    <h2>Quality Assessment</h2>
                    <div class="score">Quality Score: {data.get('quality_score', 0):.1f}/100</div>
                    
                    <div class="recommendation">
                        <strong>Recommendation:</strong> {data.get('recommendation', 'N/A')}
                    </div>
                    
                    <h2>Investment Thesis</h2>
                    <p>{data.get('thesis', 'N/A').replace(chr(10), '<br>')}</p>
                    
                    <div class="footer">
                        Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                    </div>
                </div>
            </body>
            </html>
            """
            
            with open(filename, 'w') as f:
                f.write(html_content)
            
            logger.info(f"HTML report generated: {filename}")
            return str(filename)
        except Exception as e:
            logger.error(f"Error generating HTML: {e}")
            return ""
    
    def generate_excel_report(self, symbol: str, data: Dict) -> str:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            filename = self.output_dir / f"{symbol}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis"
            
            # Title
            ws['A1'] = f"Stock Analysis Report: {symbol}"
            ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
            
            # Data
            row = 3
            for key, value in data.items():
                ws[f'A{row}'] = str(key).replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                row += 1
            
            # Auto-fit columns
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            
            wb.save(filename)
            logger.info(f"Excel report generated: {filename}")
            return str(filename)
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return ""
